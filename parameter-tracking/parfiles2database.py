import os
import pandas as pd
import yaml
from openpyxl import load_workbook
# from read-parfiles import *

def parse_parameter_file(file_path):
    """This function reads in a parameter.par file (in the seissol format) and returns all parameters in a dictionary

    Example usage:
    file_path = '/full/path/parameters.par'
    parameters = parse_parameter_file(file_path)

    Args:
        file_path (string): _description_

    Returns:
        dictionary: a dictionary with key, value pairs for all variables in the parameters.par file
    """
    parameters = {}
    
    with open(file_path, 'r') as file:
        for line in file:
            line = line.strip()
            # Skip headers and empty lines
            if line.startswith('&') or not line:
                continue
            
            # Remove comments
            if '!' in line:
                line = line.split('!')[0].strip()
            
            # Split key-value pairs
            if '=' in line:
                key, value = line.split('=')
                key = key.strip()
                value = value.strip()
                
                # Convert value to float if possible, otherwise keep as string
                try:
                    value = float(value)
                except ValueError:
                    pass
                
                parameters[key] = value
    
    return parameters

def extract_specific_parameters(parameters, keys_to_extract):
    extracted_parameters = {key: parameters[key] for key in keys_to_extract if key in parameters}
    return extracted_parameters




# # Example usage
# file_path = '/Users/hyin/agsd/projects/insar/2021_haiti/dynamic-rupture/data_tmp/scripts/parameter-tracking/inputs/parameters.par'
# parameters = parse_parameter_file(file_path)

# keys_to_extract = ['XRef', 'YRef', 'ZRef']

# extracted_parameters = extract_specific_parameters(parameters, keys_to_extract)






# # Function to read .par files
# def read_par_file(filepath):
#     parameters = {}
#     with open(filepath, 'r') as file:
#         for line in file:
#             if line.startswith("FL = "):
#                 key = 'fric_law'
#                 value = line.split('=')[1].strip().split()[0]
#                 parameters[key.strip()] = value.strip()

#             if line.startswith("MeshFile = "):
#                 key = 'meshfile'
#                 value = line.split('=')[1].strip().split()[0]
#                 parameters[key.strip()] = value.strip()

#             if line.startswith("XRef = ")
#                 value1 = line.split('=')[1].strip().split()[0]
#             if 
#     return parameters

# # Function to read .yaml files
# def read_alpha_yaml_file(filepath):

#     with open(filepath, 'r') as file:
#         parameters = yaml.safe_load(file)
#         second_alpha = data['components'][1]['components']['map']['alpha']
#     return parameters

# # Function to update the Excel file with new parameters
# def update_excel(job_id, param_dir, excel_path):
#     # List to hold all parameter dictionaries
#     all_parameters = []

#     # Loop through the directory and read each file
#     for filename in os.listdir(param_dir):
#         filepath = os.path.join(param_dir, filename)
#         if filename.endswith('.par'):
#             parameters = read_par_file(filepath)
#             parameters['filename'] = filename  # Add filename for reference
#             parameters['JobID'] = job_id       # Add JobID for reference
#             all_parameters.append(parameters)
#         elif filename.endswith('alpha.yaml'):
#             parameters = read_alpha_yaml_file(filepath)
#             parameters['filename'] = filename  # Add filename for reference
#             parameters['JobID'] = job_id       # Add JobID for reference
#             all_parameters.append(parameters)

#     # Create a DataFrame from the list of dictionaries
#     new_df = pd.DataFrame(all_parameters)

#     try:
#         # Try to load the existing workbook
#         book = load_workbook(excel_path)
#         # Load the existing sheet into a dataframe
#         with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
#             # If the Excel file already has the sheet, read it into a dataframe
#             if 'Sheet1' in book.sheetnames:
#                 existing_df = pd.read_excel(excel_path)
#                 combined_df = pd.concat([existing_df, new_df], ignore_index=True)
#                 combined_df.to_excel(writer, index=False)
#             else:
#                 # If the sheet does not exist, just write the new data
#                 new_df.to_excel(writer, index=False)
#     except FileNotFoundError:
#         # If the file does not exist, create a new one
#         new_df.to_excel(excel_path, index=False)

# # Example usage
# job_id = '12345'
# param_dir = '/dss/dsshome1/01/di35poq/haiti-rupture-inputs/dynamic-rupture/regional-only/logs/jobid_3440106'

# par_file = 'inputs/parameters.par'

# excel_path = '/hppfs/scratch/01/di35poq/haiti-rupture-outputs/test_parameters.xlsx'

# # update_excel(job_id, param_dir, excel_path)


# parameters = read_par_file(param_dir + '/' + par_file)
# print(parameters)